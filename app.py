#載入LineBot所需要的套件
from flask import Flask, request, abort, json

import schedule
import threading
import time
from datetime import datetime

import openpyxl
from openpyxl import load_workbook

import requests

from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import *

app = Flask(__name__)

access_token = 'h3GpmGFSXkxTCCsyRziMsWA6+3eCF1o5j9WSZ8os5UEDf7LqN+YqXlQFsHPdSTwijVmGMc29uutTLHuqAf28sXtFYI5VvS74M4b24bsXQaZ80fqpb0+NH65vw3ZUcrPuMqoOHWjRBDg3JCcW/dVchgdB04t89/1O/w1cDnyilFU='
channel_secret = '7befd1a844f9ea937873cd2729b04859'

# 必須放上自己的Channel Access Token
line_bot_api = LineBotApi(access_token)
# 必須放上自己的Channel Secret
handler = WebhookHandler(channel_secret)

# LINE push 訊息函式
def push_message(msg, uid, token):
    headers = {'Authorization':f'Bearer {token}','Content-Type':'application/json'}   
    body = {
    'to':uid,
    'messages':[{
            "type": "text",
            "text": msg
        }]
    }
    req = requests.request('POST', 'https://api.line.me/v2/bot/message/push', headers=headers,data=json.dumps(body).encode('utf-8'))
    print(req.text)

# LINE 回傳訊息函式
def reply_message(msg, rk, token):
    headers = {'Authorization':f'Bearer {token}','Content-Type':'application/json'}
    body = {
    'replyToken':rk,
    'messages':[{
            "type": "text",
            "text": msg
        }]
    }
    req = requests.request('POST', 'https://api.line.me/v2/bot/message/reply', headers=headers,data=json.dumps(body).encode('utf-8'))
    print(req.text)

# LINE 回傳圖片函式
def reply_image(msg, rk, token):
    headers = {'Authorization':f'Bearer {token}','Content-Type':'application/json'}
    body = {
    'replyToken':rk,
    'messages':[{
          'type': 'image',
          'originalContentUrl': msg,
          'previewImageUrl': msg
        }]
    }
    req = requests.request('POST', 'https://api.line.me/v2/bot/message/reply', headers=headers,data=json.dumps(body).encode('utf-8'))
    print(req.text)

# 儲存user_id資料
def write_user (user_id,value):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)    # 打開檔案
    s1 = wb['user_id']                                          # 開啟user_id工作表

    number = 1
    while(1):
        if s1[f'A{number}'].value is None:
            s1[f'A{number}'].value = user_id
            s1[f'B{number}'].value = value
            break
        elif s1[f'A{number}'].value == user_id:
            s1[f'B{number}'].value = value
            break
        else:
            number += 1
    
    wb.save('data.xlsx')

# 取得user_id資料
def read_user (number):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['user_id']
    read_data = s1[f'A{number}'].value

    if s1[f'B{number}'].value == 0:
        read_data = 0
        return read_data

    return read_data

# 儲存wifi帳號
def write_wifi_account(account):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)    # 打開檔案
    s1 = wb['wifi']                                          # 開啟user_id工作表

    s1['B1'].value = account
    print(account)

    wb.save('data.xlsx')

# 儲存wifi密碼
def write_wifi_password(password):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)    # 打開檔案
    s1 = wb['wifi']                                          # 開啟user_id工作表

    s1['B2'].value = password
    print(password)
    
    wb.save('data.xlsx')

# 取得wifi帳號
def read_wifi_account():
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['wifi']
    read_data = s1['B1'].value
    print(read_data)

    return read_data

# 取得wifi密碼
def read_wifi_password():
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['wifi']
    read_data = s1['B2'].value
    print(read_data)

    return read_data

# 設定定時器
def set_schelude(years, months, days, hours, minutes, user_id):
    def my_task():
        schedule.clear(user_id)
        print("任务执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        push_message('時間到該回診囉！', user_id, access_token)

    def run_scheduler():
        while True:
            schedule.run_pending()
            time.sleep(1)

    # 定义要执行任务的日期
    scheduled_date = datetime(int(years), int(months), int(days), int(hours), int(minutes))

    # 计算距离指定日期的时间差
    time_difference = scheduled_date - datetime.now()

    # 将时间差转换为秒数
    time_in_seconds = time_difference.total_seconds()

    # 定义一个定时任务，在指定日期执行
    schedule.every(time_in_seconds).seconds.do(my_task).tag(user_id)

    # 启动定时任务的线程
    scheduler_thread = threading.Thread(target=run_scheduler)
    scheduler_thread.start()


# 監聽所有來自 / 的 Post,GET Request
@app.route("/", methods=['POST','GET'])
def linebot():
    body = request.get_data(as_text=True)
    try:
        msg = request.args.get('msg')
        number = 1
        if msg == '11':
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 1，發送早安
                    push_message('早安 吃藥時間到囉! ฅ●ω●ฅ', uid=read_user(number), token=access_token)
                    number += 1
        elif msg == '21':
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 2，發送午安
                    push_message('午安 吃藥時間到囉!(๑´ㅂ`๑) ', uid=read_user(number), token=access_token)
                    number += 1
        elif msg == '31':
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 3，發送晚安
                    push_message('晚安 吃藥時間到囉! ٩(｡・ω・｡)و', uid=read_user(number), token=access_token)
                    number += 1
        elif msg == "41":
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 4，發送睡前
                    push_message('睡覺前 也該吃藥囉! (⁠⁠ꈍ⁠ᴗ⁠ꈍ⁠)', uid=read_user(number), token=access_token)
                    number += 1
        elif msg == "5":
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 4，發送睡前
                    push_message('智能藥物盒，已連接！', uid=read_user(number), token=access_token)
                    number += 1
        elif msg == "6":
            # 發送所有儲存的資料
            while(1):
                # 偵測無資料後，跳出迴圈
                if read_user(number) is None:
                    break
                elif read_user == 0:
                    number += 1
                else:
                    # 如果 msg 等於 4，發送睡前
                    push_message('智能藥物盒，沒有藥囉！', uid=read_user(number), token=access_token)
                    number += 1
        print('成功')

        signature = request.headers['X-Line-Signature']
        handler.handle(body, signature)
        json_data = json.loads(body)
        reply_token = json_data['events'][0]['replyToken']
        user_id = json_data['events'][0]['source']['userId']
        if 'message' in json_data['events'][0]:
            if json_data['events'][0]['message']['type'] == 'text':
                text = json_data['events'][0]['message']['text'].lower()
                text = text.replace('：',':').lstrip()
                print(text)
                if text == '設定:加入id':
                    write_user(user_id, 1)
                    reply_message('設定：成功設置ID', reply_token, access_token)
                elif text == '設定:移除id':
                    write_user(user_id, 0)
                    reply_message('設定：成功刪除ID', reply_token, access_token)
                elif text[:7] == '設定:回診日期':
                    date = text[7:].lstrip(' ')                                 # 單獨把日期移出來
                    year = int(date[:4].lstrip('0'))                            # 設定年
                    month = int(date[5:7].lstrip('0'))                          # 設定月
                    day = int(date[8:10].lstrip('0'))                           # 設定日
                    hour = int(date[11:13].lstrip('0'))                         # 設定時
                    if date[11:13] == '00':
                        hour = int(date[11:13].replace('0',' ',1).strip())        # 設定分
                    else:
                        hour = int(date[11:13].lstrip())
                    if date[14:16] == '00':
                        minute = int(date[14:16].replace('0',' ',1).strip())        # 設定分
                    else:
                        minute = int(date[14:16].lstrip())
                    set_schelude(year, month, day, hour, minute, user_id)
                    reply_message('設定：成功設置日期', reply_token, access_token)
                elif text[:9] == '設定:wifi帳號':
                    account = text[10:].lstrip()
                    write_wifi_account(account)
                    reply_message('設定：成功設置WiFi帳號', reply_token, access_token)
                elif text[:9] == '設定:wifi密碼':
                    password = text[10:].lstrip()
                    write_wifi_password(password)
                    reply_message('設定：成功設置WiFi密碼', reply_token, access_token)
        print('成功')
    except Exception as e:
        reply_message('設定：設置失敗', reply_token, access_token)
        print('error',e)
    return 'OK'

@app.route('/wifi')
def setup():
    data = read_wifi_account() + "," + read_wifi_password()
    return data

# 啟動應用程式
if __name__ == '__main__':
    app.run(debug=True)
